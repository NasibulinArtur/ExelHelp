import customtkinter as customtkinter
from thefuzz import fuzz
import openpyxl as openpyxl


def sim(s1,s2):
    normalized = s1.lower()
    ch = ")"
    if(ch in normalized):
        normalized1 = normalized.split(')', 1)[1]
    else: normalized1 = normalized
    normalized2 = s2.lower()
    matcher = fuzz.token_sort_ratio(normalized1,normalized2)
    return matcher


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # configure window
        self.title("Заказ поставщикам")
        customtkinter.set_appearance_mode("System")  # Modes: system (default), light, dark
        customtkinter.set_default_color_theme("dark-blue")
        self.geometry(f"{500}x{300}")

        self.button_download = customtkinter.CTkButton(self, text="Загрузить данные 1с", command=self.button_download,hover_color="#c93c20")
        self.button_download.grid(row=0, column=0, padx=10, pady=(25, 0))

        self.option_menu = customtkinter.CTkOptionMenu(self,values=["Акс", "Мва"],command=self.option_menu,dropdown_hover_color="#c93c20")
        self.option_menu.grid(row=0, column=1, padx=10, pady=(25,0))

        self.button_start = customtkinter.CTkButton(self, text="Старт", command=self.button_start, hover_color="#c93c20")
        self.button_start.grid(row=0, column=2, padx=10, pady=(25, 0))

        self.label_num_1c = customtkinter.CTkLabel(self,text="Ном. 1с",wraplength=120)
        self.label_num_1c.grid(row=1,column=0, padx=10, pady=(25,0))

        self.label_num_post = customtkinter.CTkLabel(self, text="Ном. поставщика",wraplength=110)
        self.label_num_post.grid(row=1, column=1, padx=10, pady=(25, 0))

        self.label_count= customtkinter.CTkLabel(self, text="Количество",wraplength=110)
        self.label_count.grid(row=1, column=2, padx=10, pady=(25, 0))


        self.var_continue = customtkinter.IntVar()
        self.button_to_brek = customtkinter.CTkButton(self, text="Не записывать", command=lambda: self.var_break.set(1),fg_color="#980023")
        self.button_to_brek.grid(row=3, column=0, padx=10, pady=(25, 0))

        self.var_break = customtkinter.IntVar()
        self.button_next = customtkinter.CTkButton(self, text="Записать в exel", command=lambda: self.var_continue.set(1),fg_color="#268E00")
        self.button_next.grid(row=3, column=1, padx=10, pady=(25, 0))

        self.var_next= customtkinter.IntVar()
        self.button_to_next = customtkinter.CTkButton(self, text="Следующий", command=lambda: self.var_next.set(1), hover_color="#c93c20")
        self.button_to_next.grid(row=3, column=2, padx=10, pady=(25, 0))

        self.add = customtkinter.CTkEntry(self,placeholder_text="Наименование")
        self.add.grid(row=4, column=0, padx=10, pady=(25, 0))

        self.add_count = customtkinter.CTkEntry(self, width=55,placeholder_text="кол-во")
        self.add_count.grid(row=4, column=1, padx=10, pady=(25, 0))

        self.button_add = customtkinter.CTkButton(self, text="Поиск товара", command=self.button_add, hover_color="#c93c20")
        self.button_add.grid(row=4, column=2, padx=10, pady=(25, 0))

        self.city = customtkinter.CTkEntry(self)
        self.city.grid(row=5, column=0, padx=10, pady=(25, 0))

        self.button_save = customtkinter.CTkButton(self, text="Сохранить", command=self.button_save, hover_color="#c93c20")
        self.button_save.grid(row=5, column=1, padx=10, pady=(25, 0))


    def button_download(self):
        citys = ["Уфа"]
        self.matrix = []
        self.matrix2 = []
        for city in citys:
            self.city.insert(0, city)
            # качаю ексель таблицу с данными
            self.wb1 = openpyxl.load_workbook(filename='Красноярск.xlsx')
            self.sheet1 = self.wb1.active
            for row in self.sheet1.iter_rows(min_row=13, values_only=True):
                rowlist = list(row)
                if (rowlist[13] == None):
                    rowlist[13] = 0.0
                if (rowlist[14] == None):
                    rowlist[14] = 0.0
                if (rowlist[15] == None):
                    rowlist[15] = 0.0
                if (rowlist[13] >= 1000):
                    rowlist[13] = rowlist[13] / 1000
                if (rowlist[14] >= 1000):
                    rowlist[14] = rowlist[14] / 1000
                if (rowlist[15] >= 1000):
                    rowlist[15] = rowlist[15] / 1000
                if (rowlist[14] + rowlist[15] < rowlist[13] + 3):
                    need_to_buy = rowlist[13] - rowlist[14] - rowlist[15]
                    if need_to_buy > 0:
                        self.matrix.append([rowlist[0], int(need_to_buy)])
        print(self.matrix)

    def button_start(self):

        if (self.choice == "Акс"):
            print("Это Акс")
        else :
            self.wb = openpyxl.load_workbook(filename='MVA.xlsx')
            self.sheet = self.wb["Прайс-Лист"]

        self.break_point = False
        self.con = False
        self.max = 60
        self.max_value = []

        for nom in self.matrix:
            for row in self.sheet.iter_rows(values_only=True):
                if (sim(nom[0], str(row[0])) > self.max):
                    self.max = sim(nom[0], str(row[0]))
                    try:
                        self.max_value = [str(nom[0]), row[0], sim(nom[0], str(row[0])), str(row[8]).split('*')[1], nom[1]]
                    except:
                        self.max_value = [str(nom[0]), row[0], sim(nom[0], str(row[0])), "0",nom[1]]
                if (str(row[0]) == "Процессор Intel Celeron G4930 Soc-1151v2 (3.2GHz/iUHDG610) OEM"):
                    if self.max_value != '':
                        self.matrix2.append([self.max_value])
                    self.max_value = ""
                    self.max = 60
                    break
        del self.matrix2[0]
        print(self.matrix2)
        for i in self.matrix2:
            self.label_num_1c.configure(text=i[0][0])
            self.label_num_post.configure(text=i[0][1])
            self.label_count.configure(text=str(i[0][4]))
            self.wait_variable(self.var_next)
            if (self.var_continue.get() == 1):
                self.sheet[i[0][3]].value = i[0][4]
                self.var_continue.set(0)
                print(f"Записал")
                continue
            if (self.var_break.get() == 1):
                print(f"Не записал")
                self.var_break.set(0)

        self.label_num_1c.configure(text="Конец")
        self.label_num_post.configure(text="Конец")
        self.label_count.configure(text="Конец")

    def button_add(self):
        max = 0.4
        self.add_tovar = []
        self.tovar = self.add.get()
        self.tovar_count = self.add_count.get()
        for row in self.sheet.iter_rows(values_only=True):
            if (sim(self.tovar, str(row[0])) > max):
                max = sim(self.tovar, str(row[0]))
                try:
                    self.add_tovar = [self.tovar, row[0], sim(self.tovar, str(row[0])), str(row[8]).split('*')[1], self.tovar_count]
                except:
                    self.add_tovar = ["Не найденно", "Не найденно", "Не найденно", "Не найденно", "Не найденно"]


        if(len(self.add_tovar) == 0):
            self.label_num_1c.configure(text="Пусто")
            self.label_num_post.configure(text="Пусто")
            self.label_count.configure(text="Пусто")
        else:
            self.label_num_1c.configure(text=self.add_tovar[0])
            self.label_num_post.configure(text=self.add_tovar[1])
            self.label_count.configure(text=str(self.add_tovar[4]))
            self.wait_variable(self.var_next)
            if (self.var_continue.get() == 1):
                self.sheet[self.add_tovar[3]].value = self.add_tovar[4]
                self.var_continue.set(0)
                print(f"Записал")
            if (self.var_break.get() == 1):
                print(f"Не записал")
                self.var_break.set(0)

            self.label_num_1c.configure(text="Конец")
            self.label_num_post.configure(text="Конец")
            self.label_count.configure(text="Конец")


    def option_menu(self,value):
        self.choice = value
        print(self.choice)

    def button_save(self):
        city = self.city.get()
        self.wb.save(f'{city}.xlsx')
        print("Save")



if __name__ == "__main__":
    app = App()
    app.mainloop()